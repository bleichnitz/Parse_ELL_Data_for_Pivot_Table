import openpyxl
import warnings
import filter

""" Load data from an Excel file. NOTE: the load typically throws this error "UserWarning: Data Validation
extension is not supported and will be removed warn(msg)", as openpyxl does not support data validation
(see dropdown OLG and SC columns on 'Tracking Sheet') but should NOT affect overall functioning of this program.
The 'warnings' module removes the warning from printing in the turn terminal to minimize confusion.
"""
file_name = "2022-11-18 Master Grade Summary Pivot Table.xlsx"
warnings.simplefilter(action='ignore', category=UserWarning)
print("")
print("File <" + str(file_name) + "> has loaded successfully.")
# Required Data from Master Grade Summary File:
WB_master_grade_summary = openpyxl.load_workbook(file_name)
sheetNames = WB_master_grade_summary.sheetnames
SS_master_grade_summary = WB_master_grade_summary[sheetNames[0]]
student_num = 118  # col DO
student_average = 45  # col AT
grade_level = 59  # col BH
course_code = 64  # col BM
absences = 75  # col BX
late = 89  # col CL
grade = 95  # col CR
teacher = 125  # col DV
sex = 135  # col EF
name = 136  # col EG

file_name = "2022-11-18 ELL List.xlsx"
warnings.simplefilter(action='ignore', category=UserWarning)
print("")
print("File <" + str(file_name) + "> has loaded successfully.")
WB_ell_list = openpyxl.load_workbook(file_name)
sheetNames = WB_ell_list.sheetnames
SS_ell_list = WB_ell_list[sheetNames[0]]
language_one = 16  # col Q
language_two = 17  # col R
student_number = 83  # col CF


""" 
STEP 1: clean the ESL/ELL file by removing unnecessary rows and unnecessary columns. Converts data to a working array
[[student_number, name, language_two], ...]. This data will be merged with data from the master grade summary
document.
"""
current_student = 0
previous_student = 0
ell_list = []
for row in SS_ell_list.values:
    student_data = []
    current_student = row[student_number]
    if current_student is not None:  # not a none type
        if type(current_student) is not str:  # not a string
            if int(current_student) != int(previous_student):
                student_data.append(row[student_number])
                student_data.append(row[language_one])
                student_data.append(row[language_two])
                ell_list.append(student_data)
                previous_student = current_student
                #print(student_data)
print(f"Number of ELL Students: {len(ell_list)}")

print("")
""" 
STEP 2: clean the MASTER GRADE SUMMARY file by removing unnecessary rows and unnecessary columns. Converts data to 
a working array [[student_number, name, sex, grade level, grade, average, absences, lates, course code, teacher], ...]. 
This data will be merged with data from the ESL list.
document.
"""
mgs_list = []
for row in SS_master_grade_summary.values:
    student_data = []
    current_student = row[student_num]
    ell_status = filter.filter_for_student(student=current_student, list= ell_list)
    if ell_status is not False:
        student_data.append(row[student_num])
        student_data.append(row[name])
        student_data.append(row[sex])
        student_data.append(row[grade_level])
        student_data.append(row[student_average])
        student_data.append(row[course_code])
        student_data.append(row[grade])
        student_data.append(row[absences])
        student_data.append(row[late])
        student_data.append(row[teacher])
        student_data.append(ell_status[0][1])  # language 1 from the ELL list
        student_data.append(ell_status[0][2])  # language 2 from the ELL list
        mgs_list.append([student_data])
        #print(student_data)

print(f"Number of Credits Attempted: {len(mgs_list)}")

""" 
STEP 3: write data to Pivot Table Data Doc. Create a header in row 1, and cycle
through each row to add filtered ELL data to the spread sheet so that a pivot table can be 
generated in MS Excel.
"""

pivot_table_doc = openpyxl.Workbook()
pivot_table_data_sheet = pivot_table_doc.create_sheet("Data")
#pivot_table_data_sheet.title("Data")

pivot_table_data_sheet["A1"] = "Student Number"
pivot_table_data_sheet["B1"] = "Name"
pivot_table_data_sheet["C1"] = "Sex"
pivot_table_data_sheet["D1"] = "Grade Level"
pivot_table_data_sheet["E1"] = "Average"
pivot_table_data_sheet["F1"] = "Course Code"
pivot_table_data_sheet["G1"] = "Mark"
pivot_table_data_sheet["H1"] = "Absences"
pivot_table_data_sheet["I1"] = "Lates"
pivot_table_data_sheet["J1"] = "Teacher"
pivot_table_data_sheet["K1"] = "Language 1"
pivot_table_data_sheet["L1"] = "Language 2"


for r in range(0, len(mgs_list)):
    for c in range(0, 12):
        _row = r+2
        _col = c+1
        pivot_table_data_sheet.cell(row=_row, column=_col, value=mgs_list[r][0][c])

pivot_table_doc.save("Pivot Table Data.xlsx")


# END SCRIPT
